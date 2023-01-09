import openpyxl

def convert_km_to_minutes(km, speed=30):
    hours = km / speed
    minutes = hours * 60
    return minutes

class Station:
    # Connection is a list of tuples that are consisted of (Station, time_to_reach, line)
    def __init__(self, name):
        self.name = name
        self.connections = {}
        self.distances = {}

    def add_connection(self, station, distance, line):
        if station is self:
            raise ValueError('Station to connect cannot be itself')
        self.connections[station] = {'time_distance':convert_km_to_minutes(distance), 'line': line}


    def add_time_distance(self, station, distance):
        if station is self:
            raise ValueError('Station cannot have distance to itself')
        self.distances[station] = convert_km_to_minutes(distance)
        

    def __str__(self):
        return self.name

    
class Path:
    def __init__(self, source, target):
        self.source = source
        self.current = source
        self.target = target
        self.current_line = None
        self.path = []
    
    def greedy(self):
        while(self.current != self.target):
            best_distance = 100000
            best_station = None
            print("Trying station", self.current.name)
            for station, info in self.current.connections.items():
                print(f"distance between {station.name} and {self.target.name}")
                total_distance = station.distances.get(self.target, 0)
                if total_distance < best_distance:
                    best_distance = total_distance
                    best_station = station
            if not best_station:
                raise RuntimeError("best_station should be set")
            self.path.append(best_station.name)
            self.current = best_station
            print(f"path is {self.path}")

    def a_star(self):
        current_distance = 0
        current_line = None
        while(self.current != self.target):
            best_distance = 100000
            best_station = None
            h = 4
            for end_connection in self.target.connections.values():
                for start_connection in self.current.connections.values():
                    if start_connection['line'] == end_connection['line']:
                        h = 0
                        break
             
            
            print("initial h", h)
            for station, info in self.current.connections.items():
                h = station.distances.get(self.target, 0)
                g = current_distance + self.current.distances.get(station, 0) 
                if current_line and current_line != info['line']:
                    g += 4
                current_line = info['line']
                total_distance = h + g

                if total_distance < best_distance:
                    best_distance = total_distance
                    best_station = station
            self.path.append(best_station.name)
            self.current = best_station
            current_distance += best_distance

            print("Path is", self.path)
                


        

            

        


def load_stations():
    stations = {f"e{i}": Station(f"e{i}") for i in range(1,15)}
    return stations



def load_absolute_distances(stations):
    wb = openpyxl.load_workbook('station-distances.xlsx')
    ws = wb.active
    for i, row in enumerate(ws.iter_rows()):
        for j, cell in enumerate(row): 
            if cell.value:
                try:
                    distance = float(cell.value)
                    station1 = stations[f'e{j + 1}']
                    station2 = stations[f'e{i + 1}']
                    station1.add_time_distance(station2, distance)
                    station2.add_time_distance(station1, distance)
                except ValueError:
                    pass

def load_direct_distances(stations):
    wb = openpyxl.load_workbook('direct-distances.xlsx')
    ws = wb.active
    for i, row in enumerate(ws.iter_rows()):
        for j, cell in enumerate(row):
            if cell.value:
                station1 = stations[f'e{j+1}']
                station2 = stations[f'e{i+1}']
                try:
                    distance, line = cell.value.split('-')
                    distance = float(distance.replace(',','.'))
                    station1.add_connection(station2, distance, line)            
                    station2.add_connection(station1, distance, line)  
                except ValueError:
                    pass
                




stations = load_stations()
load_absolute_distances(stations)
load_direct_distances(stations)

path = Path(stations['e1'], stations['e12'])
path.a_star()
print(path.current)


    
